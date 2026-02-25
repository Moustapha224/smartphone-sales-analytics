import sys
try:
    import openpyxl
    import json

    wb = openpyxl.load_workbook('Samsung_sales.xlsx', read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i] is not None:
                record[str(headers[i])] = val
        if any(v is not None for v in record.values()):
            data.append(record)
    wb.close()

    out = []
    out.append(f"Total rows: {len(data)}")
    out.append(f"Headers: {json.dumps(headers, ensure_ascii=False)}")
    out.append("")
    out.append("--- First 5 rows ---")
    for r in data[:5]:
        out.append(json.dumps(r, ensure_ascii=False, default=str))
    out.append("")
    out.append("--- Last 3 rows ---")
    for r in data[-3:]:
        out.append(json.dumps(r, ensure_ascii=False, default=str))

    for h in headers:
        if h is None:
            continue
        vals = sorted(set(str(r.get(str(h), '')) for r in data if r.get(str(h)) is not None))
        out.append(f"\n{h} ({len(vals)} unique): {vals[:50]}")

    out.append("\n\n--- ALL DATA JSON ---")
    out.append(json.dumps(data, ensure_ascii=False, default=str))

    with open('excel_output.txt', 'w', encoding='utf-8') as f:
        f.write('\n'.join(out))
    print("SUCCESS")
except Exception as e:
    with open('excel_error.txt', 'w', encoding='utf-8') as f:
        import traceback
        f.write(traceback.format_exc())
    print(f"ERROR: {e}")
